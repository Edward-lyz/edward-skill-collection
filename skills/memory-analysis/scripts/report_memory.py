"""Parse gpu_memory_tracker `snapshot()` log lines into a staged report.

Lines look like:
  INFO [... DP0 TP0 EP0] [GPU Memory Tracker] [after_load_target_model] driver_used=.. GB, ...

Every matching line is emitted in file order with NO de-duplication. Filter the
input to a single rank (grep 'TP0 EP0') to get one GPU's full timeline.

Sections produced:
  1. full snapshot timeline + consecutive deltas (as before)
  2. CANONICAL STAGE DELTAS -- the 12 canonical stages (initial, nccl_init,
     load_target_model, load_draft_model, alloc_target_kv_cache,
     alloc_draft_kv_cache, configure_aux_hidden,
     target_build_attention_backends, prepare_replicated_q_proj,
     draft_build_attention_backends, target_cuda_graph, draft_cuda_graph)
  3. DeepEP buffer allocations (nvl/rdma MiB)
  4. CUDA graph capture ledger (per batch size)
  5. CUDA GRAPH BREAKDOWN, one table per capture PHASE (phase = one
     cgb_total line; nested wrappers re-reporting the same capture with
     identical numbers are merged as alias classes). Buckets are
     SCOPE-EXPLICIT:
       窗口内 (sums to capture_total = the per-bs ledger sum):
         torch 图池        = private-pool segment bytes across capture()
         DeepEP NVSHMEM    = DeepEP buffer allocs inside the windows
         graphExec 实例化  = estimate: median per-window non-torch over
                             windows WITHOUT DeepEP allocs × n_windows,
                             capped at (non_torch − deepep)
         图捕获/NCCL/cubin = residual: non_torch − deepep − graphExec
       窗口外 (NOT in capture_total; why the stage delta exceeds the ledger):
         attention 图状态  = runner_init torch_alloc (graph state + static
                             buffers, logged AFTER its own capture_total and
                             re-attached to the right phase by class)
       对账 (when a before/after_*_cuda_graph snapshot pair encloses the
       phase): stage driver Δ = 窗口内 capture_total + 窗口外余量

Outputs a markdown report and, with --xlsx, an Excel workbook (snapshots,
deltas, stage_deltas, deepep_buffers, capture_ledger, cg_breakdown sheets).
"""

import re
import statistics
import sys
from typing import Dict, List, Optional, Tuple

_SNAP_RE = re.compile(
    r"\[([^\]]+)\]\s+driver_used=([\d.]+)\s*GB.*?torch_allocated=([\d.]+)\s*GB.*?"
    r"torch_reserved=([\d.]+)\s*GB.*?non_torch_used=([\d.]+)\s*GB"
    r"(?:.*?free=([\d.]+)\s*GB)?(?:.*?total=([\d.]+)\s*GB)?"
)
_DEEPEP_RE = re.compile(
    r"Allocating DeepEP buffer:\s*nvl=([\d.]+)\s*MiB,\s*rdma=([\d.]+)\s*MiB"
    r"(?:\s*\(([^)]*)\))?"
)
_LEDGER_RE = re.compile(
    r"\[capture-mem-ledger\]\s+(warmup|bs=\d+):\s+driver \+(-?[\d.]+) MiB "
    r"\(torch reserved \+(-?[\d.]+) MiB, non-torch \+(-?[\d.]+) MiB\)"
)
_CGB_INIT_RE = re.compile(
    r"\[cg-breakdown\] runner_init: torch_alloc \+(-?[\d.]+) MiB, "
    r"non_torch \+(-?[\d.]+) MiB.*?cls=(\w+)"
)
_CGB_POOL_RE = re.compile(
    r"\[cg-breakdown\] torch_graph_pool: \+(-?[\d.]+) MiB.*?cls=(\w+)"
)
_CGB_TOTAL_RE = re.compile(
    r"\[cg-breakdown\] capture_total: driver \+(-?[\d.]+) MiB, "
    r"torch_alloc \+(-?[\d.]+) MiB, torch_reserved \+(-?[\d.]+) MiB, "
    r"non_torch \+(-?[\d.]+) MiB.*?cls=(\w+)"
)

Snap = Tuple[str, float, float, float, float, float, float]

# (stage_name, before_label, after_label); "initial" is a single baseline snapshot.
CANONICAL_STAGES = [
    ("nccl_init", "before_nccl_init", "after_nccl_init"),
    ("load_target_model", "before_load_target_model", "after_load_target_model"),
    ("load_draft_model", "before_load_draft_model", "after_load_draft_model"),
    ("alloc_target_kv_cache", "before_alloc_target_kv_cache", "after_alloc_target_kv_cache"),
    ("alloc_draft_kv_cache", "before_alloc_draft_kv_cache", "after_alloc_draft_kv_cache"),
    ("configure_aux_hidden", "before_configure_aux_hidden", "after_configure_aux_hidden"),
    ("target_build_attention_backends", "before_target_build_attention_backends",
     "after_target_build_attention_backends"),
    ("prepare_replicated_q_proj", "before_prepare_replicated_q_proj",
     "after_prepare_replicated_q_proj"),
    ("draft_build_attention_backends", "before_draft_build_attention_backends",
     "after_draft_build_attention_backends"),
    ("target_cuda_graph", "before_target_cuda_graph", "after_target_cuda_graph"),
    ("draft_cuda_graph", "before_draft_cuda_graph", "after_draft_cuda_graph"),
]


def parse_events(path: str) -> List[dict]:
    """Single pass over the log; every recognized line becomes an event dict
    with 'kind' and 'lineno' so later grouping can respect file order."""
    events: List[dict] = []
    with open(path, "r", errors="ignore") as f:
        for lineno, line in enumerate(f):
            m = _SNAP_RE.search(line)
            if m and m.group(1) != "GPU Memory Tracker":
                g = m.groups()
                def _f(i):
                    return float(g[i]) if g[i] is not None else 0.0
                events.append({"kind": "snap", "lineno": lineno,
                               "snap": (m.group(1), _f(1), _f(2), _f(3),
                                        _f(4), _f(5), _f(6))})
                continue
            m = _DEEPEP_RE.search(line)
            if m:
                events.append({"kind": "deepep", "lineno": lineno,
                               "nvl": float(m.group(1)),
                               "rdma": float(m.group(2)),
                               "ctx": m.group(3) or ""})
                continue
            m = _LEDGER_RE.search(line)
            if m:
                events.append({"kind": "ledger", "lineno": lineno,
                               "stage": m.group(1),
                               "driver": float(m.group(2)),
                               "reserved": float(m.group(3)),
                               "non_torch": float(m.group(4))})
                continue
            m = _CGB_INIT_RE.search(line)
            if m:
                events.append({"kind": "cgb_init", "lineno": lineno,
                               "alloc": float(m.group(1)),
                               "non_torch": float(m.group(2)),
                               "cls": m.group(3)})
                continue
            m = _CGB_POOL_RE.search(line)
            if m:
                events.append({"kind": "cgb_pool", "lineno": lineno,
                               "mib": float(m.group(1)), "cls": m.group(2)})
                continue
            m = _CGB_TOTAL_RE.search(line)
            if m:
                events.append({"kind": "cgb_total", "lineno": lineno,
                               "driver": float(m.group(1)),
                               "alloc": float(m.group(2)),
                               "reserved": float(m.group(3)),
                               "non_torch": float(m.group(4)),
                               "cls": m.group(5)})
                continue
    return events


def snaps_of(events) -> List[Snap]:
    return [e["snap"] for e in events if e["kind"] == "snap"]


def deepep_of(events):
    return [(e["nvl"], e["rdma"], e["ctx"]) for e in events
            if e["kind"] == "deepep"]


def ledger_of(events):
    return [(e["stage"], e["driver"], e["reserved"], e["non_torch"])
            for e in events if e["kind"] == "ledger"]


def stage_deltas(snaps: List[Snap]) -> List[dict]:
    """Canonical per-stage increments. For each stage, pair each `before`
    occurrence with the NEXT `after` occurrence (multiple occurrences kept)."""
    rows: List[dict] = []
    init = next((s for s in snaps if s[0] == "initial"), None)
    if init is not None:
        rows.append({"stage": "initial", "occ": 0, "d_driver": init[1],
                     "d_alloc": init[2], "d_reserved": init[3],
                     "d_non_torch": init[4], "note": "baseline (absolute)"})
    for name, b_lbl, a_lbl in CANONICAL_STAGES:
        occ = 0
        i = 0
        while i < len(snaps):
            if snaps[i][0] == b_lbl:
                j = next((j for j in range(i + 1, len(snaps))
                          if snaps[j][0] == a_lbl), None)
                if j is None:
                    break
                b, a = snaps[i], snaps[j]
                rows.append({"stage": name, "occ": occ,
                             "d_driver": a[1] - b[1], "d_alloc": a[2] - b[2],
                             "d_reserved": a[3] - b[3],
                             "d_non_torch": a[4] - b[4], "note": ""})
                occ += 1
                i = j
            i += 1
    return rows


def cg_breakdown(events) -> List[dict]:
    """Scope-correct per-phase decomposition; see the module docstring."""
    phases: List[dict] = []
    pend_ledger: List[dict] = []
    pend_deepep: List[dict] = []
    pool_by_cls: Dict[str, float] = {}

    for e in events:
        k = e["kind"]
        if k == "ledger" and e["stage"].startswith("bs="):
            pend_ledger.append(e)
        elif k == "deepep":
            pend_deepep.append(e)
        elif k == "cgb_pool":
            pool_by_cls[e["cls"]] = e["mib"]
        elif k == "cgb_total":
            # Nested wrapper (e.g. EAGLEDraftCudaGraphRunner delegating to
            # CudaGraphRunner) re-reports the SAME capture: identical totals,
            # no windows of its own. Merge as an alias instead of emitting a
            # misleading empty table.
            if (not pend_ledger and phases
                    and abs(phases[-1]["total_driver"] - e["driver"]) < 0.05
                    and abs(phases[-1]["total_non_torch"] - e["non_torch"]) < 0.05
                    and abs(phases[-1]["total_reserved"] - e["reserved"]) < 0.05):
                phases[-1]["cls_aliases"].append(e["cls"])
                phases[-1]["lineno_close"] = e["lineno"]
                pool_by_cls.pop(e["cls"], None)
                pend_deepep = []
                continue
            # Assign each DeepEP alloc to the NEXT ledger window: the alloc
            # happens during that window's capture and its log line precedes
            # the window's ledger line.
            di = 0
            windows: List[Tuple[float, float]] = []
            for w in pend_ledger:
                dsum = 0.0
                while (di < len(pend_deepep)
                       and pend_deepep[di]["lineno"] <= w["lineno"]):
                    dsum += pend_deepep[di]["nvl"] + pend_deepep[di]["rdma"]
                    di += 1
                windows.append((max(0.0, w["non_torch"]), dsum))
            deepep_mib = (sum(d for _, d in windows)
                          + sum(x["nvl"] + x["rdma"]
                                for x in pend_deepep[di:]))
            # graphExec estimate: windows containing a DeepEP alloc would
            # smear NVSHMEM bytes into the per-graph size, so the median is
            # taken over DeepEP-free windows only; the total is capped at
            # (non_torch - deepep) so the residual stays >= 0 by construction.
            clean = [nt for nt, d in windows if d == 0.0]
            n_graphs = len(windows)
            per_graph = statistics.median(clean) if clean else None
            graph_exec = None
            capped = False
            if per_graph is not None and n_graphs:
                graph_exec = per_graph * n_graphs
                cap = max(0.0, e["non_torch"] - deepep_mib)
                capped = graph_exec > cap
                if capped:
                    graph_exec = cap
            residual = e["non_torch"] - deepep_mib - (graph_exec or 0.0)
            pool = pool_by_cls.pop(e["cls"], None)
            phases.append({
                "cls": e["cls"], "cls_aliases": [],
                "lineno_close": e["lineno"],
                "attention_graph_state": None,
                "torch_graph_pool": pool if pool is not None else e["reserved"],
                "pool_is_fallback": pool is None,
                "deepep_nvshmem": deepep_mib,
                "graph_exec": graph_exec,
                "graph_exec_per_graph": per_graph,
                "graph_exec_capped": capped,
                "capture_nccl_cubin": max(0.0, residual),
                "residual_raw": residual,
                "total_driver": e["driver"], "total_alloc": e["alloc"],
                "total_reserved": e["reserved"],
                "total_non_torch": e["non_torch"],
                "n_graphs": n_graphs,
                "stage_name": None, "stage_driver_mib": None,
                "outside_windows_mib": None,
            })
            pend_ledger, pend_deepep = [], []

    # runner_init lines are logged AFTER their own capture_total; attach each
    # to the latest phase of the same class (aliases included) closed before
    # the init line and still without an init, else the next such phase, else
    # keep it as a standalone entry so the bytes are never dropped.
    for ie in (e for e in events if e["kind"] == "cgb_init"):
        cands = [p for p in phases
                 if ie["cls"] == p["cls"] or ie["cls"] in p["cls_aliases"]]
        free = [p for p in cands if p["attention_graph_state"] is None]
        before = [p for p in free if p["lineno_close"] < ie["lineno"]]
        target = before[-1] if before else (free[0] if free else None)
        if target is not None:
            target["attention_graph_state"] = ie["alloc"]
        else:
            phases.append({
                "cls": ie["cls"], "cls_aliases": [],
                "lineno_close": ie["lineno"],
                "attention_graph_state": ie["alloc"],
                "torch_graph_pool": None, "pool_is_fallback": False,
                "deepep_nvshmem": 0.0, "graph_exec": None,
                "graph_exec_per_graph": None, "graph_exec_capped": False,
                "capture_nccl_cubin": 0.0, "residual_raw": 0.0,
                "total_driver": 0.0, "total_alloc": 0.0,
                "total_reserved": 0.0, "total_non_torch": 0.0,
                "n_graphs": 0, "stage_name": None,
                "stage_driver_mib": None, "outside_windows_mib": None,
            })

    # Reconcile with the enclosing canonical before/after_*_cuda_graph
    # snapshot pair: stage Δ = 窗口内 capture_total + 窗口外 (runner init).
    snap_evs = [e for e in events if e["kind"] == "snap"]
    pairs = []
    for i, e in enumerate(snap_evs):
        m = re.match(r"before_([a-z_]*cuda_graph)$", e["snap"][0])
        if not m:
            continue
        after = next((x for x in snap_evs[i + 1:]
                      if x["snap"][0] == "after_" + m.group(1)), None)
        if after is not None:
            pairs.append((m.group(1), e, after))
    for p in phases:
        for name, b, a in pairs:
            if b["lineno"] < p["lineno_close"] <= a["lineno"]:
                p["stage_name"] = name
                p["stage_driver_mib"] = (a["snap"][1] - b["snap"][1]) * 1024
                p["outside_windows_mib"] = (p["stage_driver_mib"]
                                            - p["total_driver"])
                break
    return phases


# 窗口内四桶：合计 = capture_total（= 逐 bs 台账之和）。
# attention_graph_state 是窗口外的 runner 初始化分配，单独一节渲染。
_WIN_BUCKETS = [
    ("torch_graph_pool", "torch 图池", "private-pool segments"),
    ("deepep_nvshmem", "DeepEP NVSHMEM", "窗口内 DeepEP allocs"),
    ("graph_exec", "graphExec 实例化",
     "est: 无 DeepEP 窗口的逐窗 non-torch 中位数 × 窗口数, 上限 non_torch−deepep"),
    ("capture_nccl_cubin", "图捕获/NCCL/cubin",
     "residual: non_torch − deepep − graphExec"),
]


def render(events) -> str:
    snaps = snaps_of(events)
    deepep = deepep_of(events)
    ledger = ledger_of(events)
    stages = stage_deltas(snaps)
    cgb = cg_breakdown(events)

    lines = ["# Staged GPU Memory Report (every stage, time order, no dedup)\n"]
    lines.append(f"{'#':>3} {'Stage':<40} {'DriverUsed':>11} {'TorchAlloc':>11} "
                 f"{'Reserved':>10} {'NonTorch':>10} {'Free':>10}")
    lines.append("-" * 100)
    for i, s in enumerate(snaps):
        lines.append(f"{i:>3} {s[0]:<40} {s[1]:>9.3f}GB {s[2]:>9.3f}GB "
                     f"{s[3]:>8.3f}GB {s[4]:>8.3f}GB {s[5]:>8.3f}GB")
    if len(snaps) >= 2:
        lines.append("\n## Consecutive deltas (driver_used / torch_allocated)")
        for i in range(1, len(snaps)):
            p, c = snaps[i - 1], snaps[i]
            lines.append(f"{i:>3} {p[0]} -> {c[0]}: driver {c[1]-p[1]:+.3f}GB, "
                         f"alloc {c[2]-p[2]:+.3f}GB, non_torch {c[4]-p[4]:+.3f}GB")

    if stages:
        lines.append("\n## Canonical stage deltas (GB)")
        lines.append(f"{'stage':<34} {'occ':>3} {'d_driver':>10} {'d_alloc':>10} "
                     f"{'d_reserved':>11} {'d_non_torch':>12}  note")
        lines.append("-" * 100)
        for r in stages:
            lines.append(f"{r['stage']:<34} {r['occ']:>3} {r['d_driver']:>+10.3f} "
                         f"{r['d_alloc']:>+10.3f} {r['d_reserved']:>+11.3f} "
                         f"{r['d_non_torch']:>+12.3f}  {r['note']}")
        present = {r["stage"] for r in stages}
        missing = [n for n, _, _ in CANONICAL_STAGES if n not in present]
        if missing:
            lines.append(f"(absent stages: {', '.join(missing)})")

    if deepep:
        lines.append("\n## DeepEP buffer allocations (non-torch)")
        lines.append(f"{'#':>3} {'nvl_MiB':>10} {'rdma_MiB':>10}  context")
        for i, (nvl, rdma, ctx) in enumerate(deepep):
            lines.append(f"{i:>3} {nvl:>10.1f} {rdma:>10.1f}  {ctx}")
        lines.append(f"TOTAL nvl={sum(d[0] for d in deepep):.1f} MiB, "
                     f"rdma={sum(d[1] for d in deepep):.1f} MiB "
                     f"({(sum(d[0] for d in deepep)+sum(d[1] for d in deepep))/1024:.3f} GB)")

    if ledger:
        lines.append("\n## CUDA graph capture ledger (per batch size, file order)")
        lines.append(f"{'#':>3} {'stage':<10} {'d_driver_MiB':>13} "
                     f"{'d_reserved_MiB':>15} {'d_non_torch_MiB':>16} {'cum_driver_MiB':>15}")
        cum = 0.0
        for i, (st, dd, dr, dn) in enumerate(ledger):
            cum += dd
            lines.append(f"{i:>3} {st:<10} {dd:>13.1f} {dr:>15.1f} "
                         f"{dn:>16.1f} {cum:>15.1f}")
        lines.append(f"TOTAL driver +{sum(l[1] for l in ledger):.1f} MiB "
                     f"(torch reserved +{sum(l[2] for l in ledger):.1f} MiB, "
                     f"non-torch +{sum(l[3] for l in ledger):.1f} MiB)")

    for wi, w in enumerate(cgb):
        cls_label = "/".join([w["cls"]] + w["cls_aliases"])
        lines.append(f"\n## CUDA graph 占用拆分 #{wi} "
                     f"(cls={cls_label}, 捕获窗口 {w['n_graphs']} 个)")
        lines.append(f"{'bucket':<44} {'MiB':>10}  note")
        lines.append("-" * 100)
        lines.append(f"[窗口内] capture_total driver +{w['total_driver']:.1f} MiB "
                     f"(torch_reserved +{w['total_reserved']:.1f}, "
                     f"non_torch +{w['total_non_torch']:.1f})"
                     f"{'' if w['n_graphs'] else '  (无独立窗口)'}")
        for key, zh, note in _WIN_BUCKETS:
            v = w[key]
            if key == "torch_graph_pool" and w["pool_is_fallback"]:
                note = "FALLBACK: reserved delta (pool stats unavailable)"
            if key == "graph_exec" and w["graph_exec_per_graph"] is not None:
                note = (f"{w['n_graphs']} 窗 × 逐窗中位数 "
                        f"{w['graph_exec_per_graph']:.1f} MiB"
                        f"（仅取无 DeepEP 分配的窗口）")
                if w["graph_exec_capped"]:
                    note += "，已按 non_torch−deepep 截断"
            if v is None:
                lines.append(f"  {key + ' (' + zh + ')':<42} {'n/a':>10}  {note}")
            else:
                lines.append(f"  {key + ' (' + zh + ')':<42} {v:>10.1f}  {note}")
        lines.append("[窗口外] runner 初始化（不计入 capture_total，也不进逐 bs 台账）:")
        v = w["attention_graph_state"]
        note = ("runner_init torch_alloc: attention 图状态 + 静态输入输出 buffer"
                if v is not None else "本相位未捕获到 runner_init 行")
        lines.append(f"  {'attention_graph_state (attention 图状态)':<42} "
                     f"{('n/a' if v is None else format(v, '.1f')):>10}  {note}")
        if w["stage_driver_mib"] is not None:
            out = w["outside_windows_mib"]
            lines.append(f"[对账] 阶段快照 {w['stage_name']}: driver Δ "
                         f"{w['stage_driver_mib']:+.1f} MiB = 窗口内 "
                         f"{w['total_driver']:+.1f} + 窗口外 {out:+.1f}"
                         + ("（窗口外≈runner 初始化张量推高的 reserved；"
                            "与 runner_init torch_alloc 的差为 slack 复用）"
                            if out > 0 else
                            "（为负：窗口内 reserved 在阶段结束前被释放/复用）"
                            if out < 0 else ""))
    return "\n".join(lines) + "\n"


def write_xlsx(events, path: str):
    from openpyxl import Workbook
    snaps = snaps_of(events)
    deepep = deepep_of(events)
    ledger = ledger_of(events)
    stages = stage_deltas(snaps)
    cgb = cg_breakdown(events)

    wb = Workbook()
    ws = wb.active
    ws.title = "snapshots"
    ws.append(["idx", "stage", "driver_used_GB", "torch_allocated_GB",
               "torch_reserved_GB", "non_torch_GB", "free_GB", "total_GB"])
    for i, s in enumerate(snaps):
        ws.append([i, s[0], round(s[1], 3), round(s[2], 3), round(s[3], 3),
                   round(s[4], 3), round(s[5], 3), round(s[6], 3)])
    wd = wb.create_sheet("deltas")
    wd.append(["idx", "from", "to", "d_driver_GB", "d_alloc_GB", "d_non_torch_GB"])
    for i in range(1, len(snaps)):
        p, c = snaps[i - 1], snaps[i]
        wd.append([i, p[0], c[0], round(c[1]-p[1], 3), round(c[2]-p[2], 3),
                   round(c[4]-p[4], 3)])
    sheets = [ws, wd]
    if stages:
        wst = wb.create_sheet("stage_deltas")
        wst.append(["stage", "occurrence", "d_driver_GB", "d_alloc_GB",
                    "d_reserved_GB", "d_non_torch_GB", "note"])
        for r in stages:
            wst.append([r["stage"], r["occ"], round(r["d_driver"], 3),
                        round(r["d_alloc"], 3), round(r["d_reserved"], 3),
                        round(r["d_non_torch"], 3), r["note"]])
        sheets.append(wst)
    if deepep:
        we = wb.create_sheet("deepep_buffers")
        we.append(["idx", "nvl_MiB", "rdma_MiB", "total_GB", "context"])
        for i, (nvl, rdma, ctx) in enumerate(deepep):
            we.append([i, round(nvl, 1), round(rdma, 1),
                       round((nvl + rdma) / 1024, 3), ctx])
        we.append(["TOTAL", round(sum(d[0] for d in deepep), 1),
                   round(sum(d[1] for d in deepep), 1),
                   round(sum(d[0] + d[1] for d in deepep) / 1024, 3), ""])
        sheets.append(we)
    if ledger:
        wl = wb.create_sheet("capture_ledger")
        wl.append(["idx", "stage", "d_driver_MiB", "d_reserved_MiB",
                   "d_non_torch_MiB", "cum_driver_MiB"])
        cum = 0.0
        for i, (st, dd, dr, dn) in enumerate(ledger):
            cum += dd
            wl.append([i, st, round(dd, 1), round(dr, 1), round(dn, 1),
                       round(cum, 1)])
        wl.append(["TOTAL", "", round(sum(l[1] for l in ledger), 1),
                   round(sum(l[2] for l in ledger), 1),
                   round(sum(l[3] for l in ledger), 1), round(cum, 1)])
        sheets.append(wl)
    if cgb:
        wc = wb.create_sheet("cg_breakdown")
        wc.append(["phase", "cls", "scope", "bucket", "bucket_zh", "MiB",
                   "n_graphs", "per_graph_MiB", "note"])
        for wi, w in enumerate(cgb):
            cls_label = "/".join([w["cls"]] + w["cls_aliases"])
            wc.append([wi, cls_label, "窗口内", "capture_total_driver",
                       "总增量(driver)", round(w["total_driver"], 1),
                       w["n_graphs"], None,
                       f"reserved={w['total_reserved']:.1f} "
                       f"non_torch={w['total_non_torch']:.1f} MiB"])
            for key, zh, note in _WIN_BUCKETS:
                v = w[key]
                ng = pg = None
                if key == "torch_graph_pool" and w["pool_is_fallback"]:
                    note = "FALLBACK: reserved delta"
                if key == "graph_exec":
                    ng = w["n_graphs"]
                    pg = (None if w["graph_exec_per_graph"] is None
                          else round(w["graph_exec_per_graph"], 1))
                    if pg is not None:
                        note = (f"{ng} 窗 × 逐窗中位数 {pg} MiB"
                                f"（仅无 DeepEP 窗口）")
                        if w["graph_exec_capped"]:
                            note += "，已截断"
                wc.append([wi, cls_label, "窗口内", key, zh,
                           None if v is None else round(v, 1), ng, pg, note])
            wc.append([wi, cls_label, "窗口外", "attention_graph_state",
                       "attention 图状态",
                       None if w["attention_graph_state"] is None
                       else round(w["attention_graph_state"], 1), None, None,
                       "runner_init torch_alloc（不计入 capture_total）"])
            if w["stage_driver_mib"] is not None:
                wc.append([wi, cls_label, "对账", "stage_driver_delta",
                           f"阶段Δ({w['stage_name']})",
                           round(w["stage_driver_mib"], 1), None, None,
                           "before/after 阶段快照差 = 窗口内 + 窗口外"])
                wc.append([wi, cls_label, "对账", "outside_windows",
                           "窗口外余量",
                           round(w["outside_windows_mib"], 1), None, None,
                           "stage Δ − capture_total"])
        sheets.append(wc)
    for w in sheets:
        for col in w.columns:
            width = max((len(str(x.value)) for x in col if x.value is not None),
                        default=8)
            w.column_dimensions[col[0].column_letter].width = min(width + 2, 44)
    wb.save(path)


def main(argv=None):
    argv = argv or sys.argv[1:]
    if not argv:
        raise SystemExit("usage: report_memory.py <tracker.log> [--xlsx out.xlsx]")
    events = parse_events(argv[0])
    if not snaps_of(events):
        raise SystemExit("no tracker snapshots found in log")
    print(render(events))
    if "--xlsx" in argv:
        out = argv[argv.index("--xlsx") + 1]
        write_xlsx(events, out)
        print(f"[written] {out}")


if __name__ == "__main__":
    main()
