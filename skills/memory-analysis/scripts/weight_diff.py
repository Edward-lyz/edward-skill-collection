"""Parse a sglang runtime weight-stats dump and diff it against theory.

The stats file is produced by `assets/weight_analysis_function.py`
(`_dump_param_memory_stats`). Its parameter table rows look like:

    model.layers.0.self_attn.q_a_proj.weight  [1536, 7168]  torch.bfloat16  11010048  21.0000

We aggregate those rows with the SAME taxonomy used for the theoretical side so
the two can be compared category by category, then explain the differences
(quantization, tensor/expert parallel sharding, fusion, tied weights).
"""

import re
from collections import defaultdict
from typing import Dict, List, Optional, Tuple

from module_taxonomy import TensorInfo, Taxonomy, CategoryAgg

_ROW_RE = re.compile(
    r"^(\S+)\s+(\[[^\]]*\])\s+(\S+)\s+(\d+)\s+([\d.]+)\s*(.*)$"
)

# torch / safetensors dtype -> canonical short tag + bytes-per-element
_DTYPE_CANON = {
    "torch.float32": ("F32", 4), "torch.float": ("F32", 4),
    "torch.bfloat16": ("BF16", 2), "torch.float16": ("F16", 2), "torch.half": ("F16", 2),
    "torch.float8_e4m3fn": ("F8_E4M3", 1), "torch.float8_e5m2": ("F8_E5M2", 1),
    "torch.int8": ("I8", 1), "torch.uint8": ("U8", 1),
    "F32": ("F32", 4), "BF16": ("BF16", 2), "F16": ("F16", 2),
    "F8_E4M3": ("F8_E4M3", 1), "F8_E5M2": ("F8_E5M2", 1), "I8": ("I8", 1), "U8": ("U8", 1),
}


def canon_dtype(d: str) -> str:
    return _DTYPE_CANON.get(d, (d, 0))[0]


def parse_stats_file(path: str) -> List[TensorInfo]:
    """Extract the parameter (and buffer) rows as measured TensorInfo."""
    tensors: List[TensorInfo] = []
    in_table = False
    with open(path, "r", errors="ignore") as f:
        for line in f:
            if line.startswith("Name ") and "Shape" in line and "Dtype" in line:
                in_table = True
                continue
            if not in_table:
                continue
            if line.startswith("=") or line.startswith("Total ") or not line.strip():
                in_table = False
                continue
            m = _ROW_RE.match(line.rstrip("\n"))
            if not m:
                continue
            name, shape_s, dtype, numel_s, mem_mb, note = m.groups()
            if "(tied)" in note or "(cpu)" in note:
                # tied/cpu rows are duplicates or off-GPU; skip from byte totals
                continue
            shape = [int(x) for x in re.findall(r"\d+", shape_s)]
            nbytes = int(round(float(mem_mb) * (1024 ** 2)))
            tensors.append(TensorInfo(name, shape, canon_dtype(dtype), nbytes))
    return tensors


def _dtype_summary(agg: CategoryAgg) -> str:
    parts = [
        f"{canon_dtype(d)}:{b / (1024 ** 3):.2f}G"
        for d, b in sorted(agg.dtypes.items(), key=lambda x: -x[1])
    ]
    return ",".join(parts)


def _module_dtype_summary(dtypes: Dict[str, int]) -> str:
    parts = [
        f"{canon_dtype(d)}:{b / (1024 ** 3):.2f}G"
        for d, b in sorted(dtypes.items(), key=lambda x: -x[1])
    ]
    return ",".join(parts)


def _module_shape_repr(shapes) -> str:
    if not shapes:
        return ""
    shown = sorted(shapes)
    s = str(list(shown[0]))
    if len(shapes) > 1:
        s += f" (+{len(shapes) - 1})"
    return s


def render_theory_report(theory: Dict[str, CategoryAgg], detailed=None,
                         by_module=None) -> str:
    # NOTE (fixed convention): every table below is sorted by NAME (category /
    # module), never by weight size. Keep it that way -- do not reintroduce
    # size-based sorting here or in write_theory_xlsx(). See
    # references/report_format.md.
    lines = ["# Theoretical Weight Sizes (from safetensors headers)\n"]

    # --- Table 1: per (category, shape, dtype), NOT aggregated over shape ---
    if detailed:
        lines.append("## By shape (category x shape x dtype)")
        lines.append(f"{'Category':<26} {'Shape':<24} {'Dtype':<10} "
                     f"{'#Tensors':>9} {'#Layers':>8} {'Total':>12}")
        lines.append("-" * 95)
        for (cat, shape, dtype), vals in sorted(
                detailed.items(), key=lambda x: (x[0][0], str(x[0][1]), x[0][2])):
            cnt, nbytes = vals[0], vals[1]
            nlayers = len(vals[2]) if len(vals) > 2 else 0
            lines.append(f"{cat:<26} {str(list(shape)):<24} {dtype:<10} "
                         f"{cnt:>9} {nlayers:>8} {nbytes / (1024 ** 3):>9.3f} GB")
        lines.append("")

    # --- Table 2: per category, aggregated, no shape column ---
    lines.append("## By category (aggregated)")
    lines.append(f"{'Category':<26} {'#Tensors':>9} {'#Layers':>8} "
                 f"{'Total':>12} {'Dtypes':<30}")
    lines.append("-" * 90)
    total = 0
    for cat in sorted(theory):
        a = theory[cat]
        total += a.total_bytes
        lines.append(f"{cat:<26} {a.count:>9} {len(a.layers):>8} "
                     f"{a.total_bytes / (1024 ** 3):>9.3f} GB {_dtype_summary(a):<30}")
    lines.append("-" * 90)
    lines.append(f"{'TOTAL':<26} {'':>9} {'':>8} {total / (1024 ** 3):>9.3f} GB")

    # --- Table 3: raw HF module names (layer/expert indices collapsed) ---
    if by_module:
        lines.append("")
        lines.append("## By module (raw HF names, layer/expert index collapsed)")
        lines.append(f"{'Module':<58} {'Category':<22} {'#Inst':>6} "
                     f"{'Shape':<22} {'PerInst':>10} {'Total':>12}")
        lines.append("-" * 134)
        mtotal = 0
        for mod in sorted(by_module):
            e = by_module[mod]
            mtotal += e["total_bytes"]
            per_inst = e["total_bytes"] / e["count"] if e["count"] else 0
            lines.append(
                f"{mod:<58} {e['category']:<22} {e['count']:>6} "
                f"{_module_shape_repr(e['shapes']):<22} "
                f"{per_inst / (1024 ** 3):>7.3f} GB {e['total_bytes'] / (1024 ** 3):>9.3f} GB")
        lines.append("-" * 134)
        lines.append(f"{'TOTAL':<58} {'':<22} {'':>6} {'':<22} {'':>10} "
                     f"{mtotal / (1024 ** 3):>9.3f} GB")

    return "\n".join(lines) + "\n"


def _autofit(ws):
    for col in ws.columns:
        width = max((len(str(x.value)) for x in col if x.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 46)


def write_theory_xlsx(theory: Dict[str, CategoryAgg], detailed, path: str,
                      by_module=None):
    from openpyxl import Workbook
    # NOTE (fixed convention): sheets sorted by NAME, never by size. See
    # render_theory_report() and references/report_format.md.
    wb = Workbook()
    ws = wb.active
    ws.title = "by_shape"
    ws.append(["category", "shape", "dtype", "num_tensors", "num_layers", "total_GB"])
    if detailed:
        for (cat, shape, dtype), vals in sorted(
                detailed.items(), key=lambda x: (x[0][0], str(x[0][1]), x[0][2])):
            nlayers = len(vals[2]) if len(vals) > 2 else 0
            ws.append([cat, str(list(shape)), dtype, vals[0], nlayers,
                       round(vals[1] / (1024 ** 3), 4)])
    wc = wb.create_sheet("by_category")
    wc.append(["category", "num_tensors", "num_layers", "total_GB", "dtypes"])
    for cat in sorted(theory):
        a = theory[cat]
        wc.append([cat, a.count, len(a.layers), round(a.total_bytes / (1024 ** 3), 4),
                   _dtype_summary(a)])
    if by_module:
        wm = wb.create_sheet("by_module")
        wm.append(["module", "category", "num_instances", "num_layers", "shape",
                   "dtypes", "per_instance_GB", "total_GB"])
        for mod in sorted(by_module):
            e = by_module[mod]
            per_inst = e["total_bytes"] / e["count"] if e["count"] else 0
            wm.append([mod, e["category"], e["count"], len(e["layers"]),
                       _module_shape_repr(e["shapes"]),
                       _module_dtype_summary(e["dtypes"]),
                       round(per_inst / (1024 ** 3), 4),
                       round(e["total_bytes"] / (1024 ** 3), 4)])
        _autofit(wm)
    _autofit(ws); _autofit(wc)
    wb.save(path)


def write_diff_xlsx(theory, measured, mapping_rows, cfg, path: str):
    from openpyxl import Workbook
    entry_by_cat = {r.category: r for r in mapping_rows}
    wb = Workbook()
    ws = wb.active
    ws.title = "diff"
    ws.append(["category", "theory_GB", "expect_GB(config)", "measured_GB",
               "delta_GB", "obs_ratio", "axis", "note"])
    for cat in sorted(set(theory) | set(measured)):
        t = theory.get(cat)
        mo = measured.get(cat)
        row = entry_by_cat.get(cat)
        axis = row.tp if row else "replicated"
        theory_gb = t.total_bytes / (1024 ** 3) if t else 0.0
        meas_gb = mo.total_bytes / (1024 ** 3) if mo else 0.0
        obs = (t.total_bytes / mo.total_bytes) if (t and mo and mo.total_bytes) else None
        deg = axis_degree(axis, cfg) if (cfg and t) else None
        expect_gb = (t.total_bytes / deg / (1024 ** 3)) if (deg and t) else None
        note = ""
        if t and not mo:
            note = "only in HF (fused/renamed?)"
        elif mo and not t:
            note = "only in sglang (buffer/derived?)"
        else:
            td = set(canon_dtype(d) for d in t.dtypes)
            md = set(canon_dtype(d) for d in mo.dtypes)
            if td != md:
                note = f"dtype {sorted(td)}->{sorted(md)}"
            elif cfg and expect_gb is not None and abs(meas_gb - expect_gb) > max(
                    0.05 * expect_gb, 0.05):
                note = "MISMATCH vs config"
        ws.append([cat, round(theory_gb, 4),
                   round(expect_gb, 4) if expect_gb is not None else None,
                   round(meas_gb, 4),
                   round(meas_gb - expect_gb, 4) if (expect_gb is not None and t and mo) else None,
                   round(obs, 2) if obs else None, axis, note])
    wm = wb.create_sheet("mapping")
    wm.append(["category", "sglang_param", "fusion", "axis", "note"])
    for r in mapping_rows:
        wm.append([r.category, r.sglang_hint, r.fusion, r.tp, r.note])
    _autofit(ws); _autofit(wm)
    wb.save(path)


def infer_tp(theory: Dict[str, CategoryAgg],
             measured: Dict[str, CategoryAgg]) -> int:
    """Infer the TP/EP degree from theory-vs-measured byte ratios.

    Sharded weights show `theory/measured ~= TP` while replicated ones show ~1.
    We look at significant categories present on both sides with matching dtype
    (so quantization doesn't skew the ratio), round each ratio, and take the most
    common value among the sharded (>=1.5) ones. Returns 1 if nothing looks
    sharded (single-rank dump).
    """
    from collections import Counter
    sharded = []
    for cat in set(theory) & set(measured):
        t = theory[cat].total_bytes
        m = measured[cat].total_bytes
        if t < 512 * 1024 * 1024 or m <= 0:
            continue
        td = set(canon_dtype(d) for d in theory[cat].dtypes)
        md = set(canon_dtype(d) for d in measured[cat].dtypes)
        if td != md:
            continue
        r = t / m
        if r >= 1.5:
            sharded.append(round(r))
    if not sharded:
        return 1
    return Counter(sharded).most_common(1)[0][0]


# Logical shard axis -> config degree. The axis is declared by the model adapter
# (a STRUCTURAL fact about how sglang shards that category); the numeric degree
# comes from the launch config. Never derive the degree from the measured dump.
_AXIS_ALIAS = {"col_parallel": "tensor", "row_parallel": "tensor",
               "ep": "moe_ep", "tp": "tensor", "dcp": "attn_tp"}


def axis_degree(axis: str, cfg: Dict[str, int]):
    """Resolve a category's expected shard degree from config, or None if the
    axis is unknown (cannot predict -> reported as '?', never fitted)."""
    a = _AXIS_ALIAS.get(axis, axis)
    if a == "replicated":
        return 1
    if a == "moe_ep":
        return max(cfg.get("ep", 1), 1)
    if a == "attn_tp":
        d = cfg.get("dcp", 1)
        return d if d > 1 else max(cfg.get("tp", 1), 1)
    if a == "tensor":
        return max(cfg.get("tp", 1), 1)
    return None


def render_diff_report(theory: Dict[str, CategoryAgg],
                       measured: Dict[str, CategoryAgg],
                       mapping_rows,
                       cfg: Optional[Dict[str, int]],
                       stacked_scan: Dict[str, List[str]]) -> str:
    entry_by_cat = {r.category: r for r in mapping_rows}
    lines = ["# HF Theory <-> sglang Measured Diff\n"]
    if cfg:
        lines.append(f"Parallel config (from launch, independent of the dump): "
                     f"tp={cfg.get('tp',1)} ep={cfg.get('ep',1)} "
                     f"dcp={cfg.get('dcp',1)} dp={cfg.get('dp',1)}")
        lines.append("`Expect` = Theory / degree(config). `ObsRatio` = Theory / "
                     "Measured (display only). Δ = Measured - Expect. Degrees come "
                     "from config, NOT from the dump, so Δ is a real check.\n")
    else:
        lines.append("NO parallel config supplied: `Expect`/Δ omitted on purpose "
                     "(predicting shard degree from the dump would be circular). "
                     "Showing `ObsRatio` = Theory/Measured only. Pass "
                     "--yaml/--tp-size/--ep-size/--dcp-size to enable the check.\n")

    # --- mapping table ---
    lines.append("## HF category -> sglang runtime mapping")
    lines.append(f"{'Category':<26} {'sglang param':<34} {'fusion':<8} {'axis':<12} note")
    lines.append("-" * 100)
    for r in mapping_rows:
        lines.append(f"{r.category:<26} {r.sglang_hint:<34} {r.fusion:<8} {r.tp:<12} {r.note}")
    if stacked_scan:
        lines.append("\n### sglang stacked_params_mapping found in source (ground truth)")
        for fused, srcs in sorted(stacked_scan.items()):
            lines.append(f"  {fused} <- {', '.join(srcs)}")

    # --- per-category diff ---
    lines.append("\n## Per-category theory vs measured (per rank)")
    if cfg:
        lines.append(f"{'Category':<26} {'Theory':>10} {'Expect':>10} {'Measured':>10} "
                     f"{'Delta':>10} {'ObsRatio':>9}  note")
    else:
        lines.append(f"{'Category':<26} {'Theory':>10} {'Measured':>10} "
                     f"{'ObsRatio':>9}  note")
    lines.append("-" * 108)
    suggestions: List[str] = []
    for cat in sorted(set(theory) | set(measured)):
        t = theory.get(cat)
        mo = measured.get(cat)
        row = entry_by_cat.get(cat)
        axis = row.tp if row else "replicated"
        theory_gb = t.total_bytes / (1024 ** 3) if t else 0.0
        meas_gb = mo.total_bytes / (1024 ** 3) if mo else 0.0
        obs = (t.total_bytes / mo.total_bytes) if (t and mo and mo.total_bytes) else None
        obs_s = f"{obs:>8.1f}x" if obs else "     -"
        deg = axis_degree(axis, cfg) if (cfg and t) else None
        expect_gb = (t.total_bytes / deg / (1024 ** 3)) if (deg and t) else None

        note = ""
        if t and not mo:
            note = "only in HF (fused/renamed in sglang?)"
        elif mo and not t:
            note = "only in sglang (buffer/derived?)"
        else:
            td = set(canon_dtype(d) for d in t.dtypes)
            md = set(canon_dtype(d) for d in mo.dtypes)
            if td != md:
                note = f"dtype changed {sorted(td)}->{sorted(md)}"
                if (not cat.endswith(".__quant__")
                        and any(x.startswith("F8") for x in md)
                        and not any(x.startswith("F8") for x in td)):
                    suggestions.append(
                        f"- {cat}: HF is {sorted(td)} but runtime is fp8 -> "
                        f"quantization active; expected saving already realised.")
            elif cfg and expect_gb is not None:
                delta = meas_gb - expect_gb
                if abs(delta) > 0.05 * max(expect_gb, 1e-9) and abs(delta) > 0.05:
                    note = "MISMATCH vs config prediction (real finding)"
            elif cfg and deg is None:
                note = "unknown axis: cannot predict"

        if cfg:
            exp_s = f"{expect_gb:>8.3f}GB" if expect_gb is not None else "        ?"
            dlt_s = (f"{meas_gb - expect_gb:>+8.3f}GB"
                     if (expect_gb is not None and t and mo) else "        -")
            lines.append(f"{cat:<26} {theory_gb:>8.3f}GB {exp_s} {meas_gb:>8.3f}GB "
                         f"{dlt_s} {obs_s:>9}  {note}")
        else:
            lines.append(f"{cat:<26} {theory_gb:>8.3f}GB {meas_gb:>8.3f}GB "
                         f"{obs_s:>9}  {note}")

    # --- heuristic suggestions ---
    for cat, t in theory.items():
        if cat.endswith(".__quant__"):
            continue
        dts = set(canon_dtype(d) for d in t.dtypes)
        mo = measured.get(cat)
        md = set(canon_dtype(d) for d in mo.dtypes) if mo else set()
        big = t.total_bytes / (1024 ** 3) > 1.0
        if big and dts <= {"BF16", "F16"} and not any(x.startswith("F8") for x in md):
            suggestions.append(
                f"- {cat}: {t.total_bytes / (1024 ** 3):.1f}GB still 16-bit at "
                f"runtime; candidate for fp8/int quantization.")
    lines.append("\n## Optimization suggestions")
    if suggestions:
        lines.extend(sorted(set(suggestions)))
    else:
        lines.append("- No obvious weight-level optimization found from this diff.")
    return "\n".join(lines) + "\n"
